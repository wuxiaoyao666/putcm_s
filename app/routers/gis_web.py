import io
import os
import re
from typing import Dict

from fastapi import APIRouter, UploadFile, File, Depends
from openpyxl import load_workbook
from sqlalchemy import select, tuple_
from sqlalchemy.ext.asyncio import AsyncSession
from starlette.responses import FileResponse

from app import User
from app.database import get_db
from app.models.gis import Gis
from app.models.tcm import Tcm
from app.models.user import get_curr_user

router = APIRouter(prefix="/api/gis", tags=["gis"])

# Excel 表头与数据库字段的映射关系
HEADER_MAPPING = {
    "药材名称": "tcmName",
    "样品类型": "sampleType",
    "采样类型": "samplingType",
    "样地编号": "plotNumber",
    "样品编号": "sampleNumber",
    "省份": "province",
    "城市": "city",
    "区县": "district",
    "乡镇": "township",
    "村庄": "village",
    "经度": "lng",
    "纬度": "lat",
    "海拔": "altitude",
    "采样人": "collector",
    "采样单位": "collectionUnit",
    "备注": "remarks"
}


@router.get("/template", summary="下载GIS导入模板")
async def download_template():
    # 文件的相对路径
    file_path = "file/template/gis.xlsx"

    # 检查文件是否存在
    if not os.path.exists(file_path):
        return {"status": 1, "msg": "服务器端模板文件丢失，请联系管理员"}

    # 定义下载时的文件名
    download_filename = "GIS数据导入模板.xlsx"

    # 直接返回文件响应
    return FileResponse(
        path=file_path,
        filename=download_filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.post("/upload", summary="上传Excel导入GIS数据")
async def upload_excel(
        file: UploadFile = File(...),
        curr_user: User = Depends(get_curr_user),
        db: AsyncSession = Depends(get_db)
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        return {"status": 1, "msg": "请上传 Excel 文件"}

    try:
        # 读取 Excel 文件
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active

        # 获取表头并验证
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {"status": 1, "msg": "文件内容为空"}

        headers = rows[0]
        header_map = {}

        for idx, header in enumerate(headers):
            if header in HEADER_MAPPING:
                header_map[idx] = HEADER_MAPPING[header]

        if "tcmName" not in header_map.values():
            return {"status": 1, "msg": "模板错误：缺少'药材名称'列"}

        # 预加载所有中药名称与ID的映射
        tcm_query = select(Tcm.tcmName, Tcm.tcmId)
        tcm_res = await db.execute(tcm_query)
        tcm_map: Dict[str, int] = {name: tcm_id for name, tcm_id in tcm_res.all()}

        # 暂存解析后的行
        parsed_rows = []
        errors = []

        # 1. 解析所有 Excel 行数据
        for row_idx, row in enumerate(rows[1:], start=2):
            row_data = {}
            tcm_name = None

            for col_idx, value in enumerate(row):
                if col_idx in header_map:
                    field_name = header_map[col_idx]
                    if field_name == "tcmName":
                        tcm_name = str(value).strip() if value else None
                    else:
                        row_data[field_name] = str(value).strip() if value is not None else None

            if not tcm_name:
                continue
            if tcm_name not in tcm_map:
                errors.append(f"第{row_idx}行：药材'{tcm_name}'不存在")
                continue

            row_data["tcmId"] = tcm_map[tcm_name]

            # 唯一性判断
            if not row_data.get('plotNumber') or not row_data.get('sampleNumber'):
                errors.append(f"第{row_idx}行：样地编号或样品编号缺失，无法处理")
                continue

            parsed_rows.append(row_data)

        if errors:
            return {"status": 1, "msg": "导入失败", "errors": errors[:5]}
        if not parsed_rows:
            return {"status": 1, "msg": "没有有效数据可导入"}

        # 2. Excel 内部去重 key: (plotNumber, sampleNumber) -> value: row_data
        unique_data_map = {}
        for row in parsed_rows:
            key = (row['plotNumber'], row['sampleNumber'])
            unique_data_map[key] = row

        final_rows = list(unique_data_map.values())
        keys_to_check = list(unique_data_map.keys())

        # 3. 查询数据库中已存在的记录
        existing_map = {}  # (plotNumber, sampleNumber) -> subId

        if keys_to_check:
            stmt = select(Gis.subId, Gis.plotNumber, Gis.sampleNumber).where(
                tuple_(Gis.plotNumber, Gis.sampleNumber).in_(keys_to_check)
            )
            existing_res = await db.execute(stmt)
            for r in existing_res.all():
                existing_map[(r.plotNumber, r.sampleNumber)] = r.subId

        # 4. 分流：区分 新增列表 和 更新列表
        to_insert = []
        to_update = []

        for row in final_rows:
            key = (row['plotNumber'], row['sampleNumber'])
            if key in existing_map:
                # 存在：添加主键 subId，放入更新列表
                row['subId'] = existing_map[key]
                to_update.append(row)
            else:
                # 不存在：放入新增列表
                to_insert.append(row)

        # 5. 执行批量操作
        if to_insert:
            await Gis.bulk_add(db, to_insert, curr_user.userId)

        if to_update:
            await Gis.bulk_update(db, to_update, curr_user.userId)

        msg = []
        if to_insert:
            msg.append(f"新增 {len(to_insert)} 条")
        if to_update:
            msg.append(f"更新 {len(to_update)} 条")
        # 提交事务
        await db.commit()

        return {"status": 0, "msg": "，".join(msg) + " 数据已处理成功"}

    except Exception as e:
        # 回滚
        await db.rollback()
        import traceback
        traceback.print_exc()
        return {"status": 1, "msg": f"处理文件时发生错误: {str(e)}"}


@router.get("/list")
async def get_list(tcmId: int, db: AsyncSession = Depends(get_db)):
    data = await Gis.get_list(db, tcmId)
    return {"status": 0, "data": data}
